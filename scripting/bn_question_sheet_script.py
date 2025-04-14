from openpyxl import Workbook, load_workbook
from shared_utils.file_service import open_txt


da_wb = load_workbook("./outputs/sampling/sampled_translations.xlsx")
da_ws = da_wb.active



english_qsts = open_txt("./outputs/questions_txt_files/dev_questions.txt")
qst_index_map = {}
for index, qst in enumerate(english_qsts):
    qst_index_map[qst] = index

bengali_qsts = open_txt("./outputs/translations/google/dev_questions_bn_linebyline.txt")

questions_for_sampling = []

print(english_qsts[1])
for row in da_ws.iter_rows(min_row=2, max_row=da_ws.max_row, min_col=1, max_col=2, values_only=True):
    print(row)
    eng_qst = row[0]
    index = qst_index_map[eng_qst]
    bn_qst = bengali_qsts[index]
    questions_for_sampling.append([eng_qst,bn_qst])


bn_wb = Workbook()
bn_ws = bn_wb.active
bn_ws.append(["English", "Bengali"])
for qsts in questions_for_sampling:
    bn_ws.append(qsts)

bn_wb.save("./outputs/sampling/sampled_translations_bn.xlsx")



